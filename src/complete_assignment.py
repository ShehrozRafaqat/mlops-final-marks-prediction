"""Complete Assignment 02 deliverables.

This script trains predictive models from the morning sections and predicts
each afternoon student's final total after the 5th activity through the
second-last activity. It generates:

1. outputs/Assignment_02_Predictions.xlsx
2. outputs/Assignment_02_Report.pdf
"""

from __future__ import annotations

import argparse
import textwrap
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.base import clone
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import ElasticNet, Ridge
from sklearn.metrics import mean_absolute_error
from sklearn.model_selection import KFold
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler


START_HORIZON = 5


@dataclass(frozen=True)
class CourseConfig:
    course_name: str
    file_name: str
    train_sheet: str
    test_sheet: str
    target_col: int
    first_activity_col: int = 3
    activity_header_row: int = 1
    weights_row: int = 0
    scale_row: int = 2
    data_start_row: int = 3
    id_col: int = 1
    model_strategy: str = "ridge_remaining_ratio"
    dropout_start_horizon: int = 8
    dropout_earned_threshold: float = 3.0
    dropout_missing_zero_threshold: float = 0.5


@dataclass
class CourseDataset:
    config: CourseConfig
    student_ids: pd.Series
    features: pd.DataFrame
    target: pd.Series
    weights: pd.Series
    scales: pd.Series
    observed: pd.DataFrame


COURSES = [
    CourseConfig(
        course_name="Cloud Computing",
        file_name="[Template] CC Result Data Set.xlsx",
        train_sheet="cloud computing morning",
        test_sheet="cloud computing afternoon",
        target_col=17,
        model_strategy="random_forest_remaining_ratio",
        dropout_start_horizon=7,
        dropout_earned_threshold=3.0,
        dropout_missing_zero_threshold=0.9,
    ),
    CourseConfig(
        course_name="ICT",
        file_name="[Template] ICT Result Data Set.xlsx",
        train_sheet="ICT Morning",
        test_sheet="ICT Afternoon",
        target_col=15,
        model_strategy="elastic_net_remaining_ratio",
        dropout_start_horizon=5,
        dropout_earned_threshold=3.0,
        dropout_missing_zero_threshold=0.4,
    ),
]


def ordinal(number: int) -> str:
    if 10 <= number % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(number % 10, "th")
    return f"{number}{suffix}"


def make_unique(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique = []
    for name in names:
        count = seen.get(name, 0)
        unique.append(name if count == 0 else f"{name}_{count + 1}")
        seen[name] = count + 1
    return unique


def load_course_dataset(data_dir: Path, config: CourseConfig, sheet_name: str) -> CourseDataset:
    workbook_path = data_dir / config.file_name
    raw = pd.read_excel(workbook_path, sheet_name=sheet_name, header=None)

    activity_slice = slice(config.first_activity_col, config.target_col)
    activity_names = raw.iloc[config.activity_header_row, activity_slice].tolist()
    activity_names = [
        str(value).strip() if pd.notna(value) and str(value).strip() else f"Activity {idx + 1}"
        for idx, value in enumerate(activity_names)
    ]
    activity_names = make_unique(activity_names)

    weights = pd.to_numeric(raw.iloc[config.weights_row, activity_slice], errors="coerce")
    scales = pd.to_numeric(raw.iloc[config.scale_row, activity_slice], errors="coerce")
    weights = pd.Series(weights.to_numpy(dtype=float), index=activity_names).fillna(0.0)
    scales = pd.Series(scales.to_numpy(dtype=float), index=activity_names).replace(0, np.nan).fillna(1.0)

    data = raw.iloc[config.data_start_row :].copy()
    data = data[data.iloc[:, config.id_col].notna()].reset_index(drop=True)

    student_ids = data.iloc[:, config.id_col].astype(str).str.strip()
    target = pd.to_numeric(data.iloc[:, config.target_col], errors="coerce")

    features = data.iloc[:, config.first_activity_col : config.target_col].copy()
    features.columns = activity_names
    features = features.apply(pd.to_numeric, errors="coerce")
    observed = features.notna()
    features = features.fillna(0.0)

    valid_rows = target.notna()
    return CourseDataset(
        config=config,
        student_ids=student_ids.loc[valid_rows].reset_index(drop=True),
        features=features.loc[valid_rows].reset_index(drop=True),
        target=target.loc[valid_rows].astype(float).reset_index(drop=True),
        weights=weights,
        scales=scales,
        observed=observed.loc[valid_rows].reset_index(drop=True),
    )


def engineered_features(dataset: CourseDataset, horizon: int) -> pd.DataFrame:
    known_names = list(dataset.features.columns[:horizon])
    raw = dataset.features.loc[:, known_names].copy()
    observed = dataset.observed.loc[:, known_names].astype(float)

    scales = dataset.scales.loc[known_names]
    weights = dataset.weights.loc[known_names]
    normalized = raw.divide(scales, axis="columns").clip(lower=0.0)
    contribution = normalized.multiply(weights, axis="columns")

    feature_blocks = [
        raw.add_prefix("raw__"),
        normalized.add_prefix("ratio__"),
        contribution.add_prefix("earned__"),
        observed.add_prefix("observed__"),
    ]
    features = pd.concat(feature_blocks, axis=1)

    total_weight = float(weights.sum())
    cumulative = contribution.sum(axis=1)
    features["cumulative_earned"] = cumulative
    features["known_weight"] = total_weight
    features["known_weight_ratio"] = total_weight / float(dataset.weights.sum())
    features["earned_ratio_so_far"] = cumulative / total_weight if total_weight else 0.0
    features["observed_count"] = observed.sum(axis=1)
    features["observed_ratio"] = observed.mean(axis=1)
    return features


def cumulative_earned(dataset: CourseDataset, horizon: int) -> pd.Series:
    known_names = list(dataset.features.columns[:horizon])
    raw = dataset.features.loc[:, known_names]
    contribution = raw.divide(dataset.scales.loc[known_names], axis="columns").multiply(
        dataset.weights.loc[known_names], axis="columns"
    )
    return contribution.sum(axis=1)


def model_for_course(config: CourseConfig, random_state: int) -> tuple[str, object]:
    if config.model_strategy == "random_forest_remaining_ratio":
        return (
            "Remaining-Ratio Random Forest",
            RandomForestRegressor(
                n_estimators=120,
                min_samples_leaf=3,
                random_state=random_state,
                n_jobs=-1,
            ),
        )

    if config.model_strategy == "elastic_net_remaining_ratio":
        return (
            "Remaining-Ratio Elastic Net",
            make_pipeline(
                StandardScaler(),
                ElasticNet(alpha=0.1, l1_ratio=0.2, max_iter=5000, random_state=random_state),
            ),
        )

    return (
        "Remaining-Ratio Ridge Regression",
        make_pipeline(StandardScaler(), Ridge(alpha=10.0)),
    )


def remaining_ratio_target(dataset: CourseDataset, horizon: int) -> pd.Series:
    remaining_weight = max(float(dataset.weights.iloc[horizon:].sum()), 1e-9)
    remaining_marks = (dataset.target - cumulative_earned(dataset, horizon)).clip(lower=0.0)
    return remaining_marks / remaining_weight


def predict_remaining_ratio(
    model: object,
    train: CourseDataset,
    predict_dataset: CourseDataset,
    horizon: int,
) -> np.ndarray:
    x_train = engineered_features(train, horizon)
    y_train = remaining_ratio_target(train, horizon)
    x_predict = engineered_features(predict_dataset, horizon)
    known_marks = cumulative_earned(predict_dataset, horizon).to_numpy(dtype=float)
    remaining_weight = float(predict_dataset.weights.iloc[horizon:].sum())

    fitted_model = clone(model)
    fitted_model.fit(x_train, y_train)
    remaining_ratio = fitted_model.predict(x_predict)
    predicted = known_marks + remaining_ratio * remaining_weight
    return np.clip(predicted, known_marks, 100.0)


def training_cv_mae_for_strategy(
    model: object,
    dataset: CourseDataset,
    horizon: int,
    random_state: int,
) -> float:
    x_train = engineered_features(dataset, horizon)
    y_train = remaining_ratio_target(dataset, horizon)
    cv = KFold(n_splits=min(5, len(y_train)), shuffle=True, random_state=random_state)
    remaining_weight = float(dataset.weights.iloc[horizon:].sum())
    known_marks = cumulative_earned(dataset, horizon).to_numpy(dtype=float)

    errors = []
    for train_idx, validation_idx in cv.split(x_train):
        fitted_model = clone(model)
        fitted_model.fit(x_train.iloc[train_idx], y_train.iloc[train_idx])
        remaining_ratio = fitted_model.predict(x_train.iloc[validation_idx])
        validation_known = known_marks[validation_idx]
        predicted = validation_known + remaining_ratio * remaining_weight
        predicted = np.clip(predicted, validation_known, 100.0)
        errors.append(mean_absolute_error(dataset.target.iloc[validation_idx], predicted))

    return float(np.mean(errors))


def disengaged_students(dataset: CourseDataset, horizon: int) -> pd.Series:
    config = dataset.config
    if horizon < config.dropout_start_horizon:
        return pd.Series(False, index=dataset.features.index)

    known_names = list(dataset.features.columns[:horizon])
    known_marks = cumulative_earned(dataset, horizon)
    raw = dataset.features.loc[:, known_names]
    missing_or_zero_ratio = ((~dataset.observed.loc[:, known_names]) | (raw <= 0)).mean(axis=1)

    return (
        (known_marks <= config.dropout_earned_threshold)
        & (missing_or_zero_ratio >= config.dropout_missing_zero_threshold)
    )


def prediction_column_name(horizon: int, activity_name: str, total_activities: int) -> str:
    label = f"Predicted after {ordinal(horizon)} ({activity_name})"
    if horizon == total_activities - 1:
        label += " - Second Last"
    return label


def predict_course(
    train: CourseDataset,
    test: CourseDataset,
    random_state: int,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    total_activities = len(train.features.columns)
    prediction_horizons = range(START_HORIZON, total_activities)

    predictions = pd.DataFrame(
        {
            "Student ID": test.student_ids,
            "Actual Score": test.target.round(2),
        }
    )
    model_records = []

    for horizon in prediction_horizons:
        model_name, model = model_for_course(train.config, random_state + horizon)
        cv_mae = training_cv_mae_for_strategy(model, train, horizon, random_state + horizon)
        predicted = predict_remaining_ratio(model, train, test, horizon)

        disengaged_mask = disengaged_students(test, horizon).to_numpy()
        known_marks = cumulative_earned(test, horizon).to_numpy(dtype=float)
        predicted = np.where(disengaged_mask, known_marks, predicted)
        full_model_name = f"{model_name} + engagement fallback"

        activity_name = str(train.features.columns[horizon - 1])
        column_name = prediction_column_name(horizon, activity_name, total_activities)
        predictions[column_name] = np.round(predicted, 2)
        model_records.append(
            {
                "Course": train.config.course_name,
                "Horizon": horizon,
                "Activity": activity_name,
                "Selected Model": full_model_name,
                "Training CV MAE": round(cv_mae, 3),
                "Test MAE": round(mean_absolute_error(test.target, predicted), 3),
                "Fallback Count": int(disengaged_mask.sum()),
            }
        )

    pred_cols = [col for col in predictions.columns if col.startswith("Predicted after")]
    absolute_errors = predictions[pred_cols].sub(predictions["Actual Score"], axis=0).abs()
    metrics = pd.DataFrame(
        {
            "Student ID": predictions["Student ID"],
            "Average Absolute Error": absolute_errors.mean(axis=1).round(2),
            "Max Absolute Error": absolute_errors.max(axis=1).round(2),
            "Min Absolute Error": absolute_errors.min(axis=1).round(2),
        }
    )
    average_row = {
        "Student ID": "AVERAGE",
        "Average Absolute Error": round(float(metrics["Average Absolute Error"].mean()), 2),
        "Max Absolute Error": round(float(metrics["Max Absolute Error"].mean()), 2),
        "Min Absolute Error": round(float(metrics["Min Absolute Error"].mean()), 2),
    }
    metrics = pd.concat([metrics, pd.DataFrame([average_row])], ignore_index=True)

    return predictions, metrics, pd.DataFrame(model_records)


def autosize_excel(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    average_fill = PatternFill("solid", fgColor="FFF2CC")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "C2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

        for row in worksheet.iter_rows(min_row=2):
            if row[0].value == "AVERAGE":
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = average_fill

        for column_cells in worksheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 28)

    workbook.save(path)


def write_prediction_workbook(path: Path, results: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for course_name, predictions in results.items():
            predictions.to_excel(writer, sheet_name=course_name[:31], index=False)
    autosize_excel(path)


def add_text_page(pdf: PdfPages, title: str, paragraphs: list[str]) -> None:
    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor("white")
    fig.text(0.08, 0.94, title, fontsize=18, weight="bold", va="top")

    y = 0.88
    for paragraph in paragraphs:
        wrapped = textwrap.fill(paragraph, width=92)
        fig.text(0.08, y, wrapped, fontsize=10.5, va="top", linespacing=1.35)
        y -= 0.04 * (wrapped.count("\n") + 2)

    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def add_dataframe_table(
    pdf: PdfPages,
    title: str,
    dataframe: pd.DataFrame,
    rows_per_page: int = 28,
) -> None:
    chunks = [
        dataframe.iloc[start : start + rows_per_page]
        for start in range(0, len(dataframe), rows_per_page)
    ]

    for page_index, chunk in enumerate(chunks, start=1):
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis("off")
        page_title = title if len(chunks) == 1 else f"{title} ({page_index}/{len(chunks)})"
        ax.set_title(page_title, fontsize=14, weight="bold", pad=14)

        formatted = chunk.copy()
        for col in formatted.columns:
            if pd.api.types.is_numeric_dtype(formatted[col]):
                formatted[col] = formatted[col].map(lambda value: f"{float(value):.2f}")
            else:
                formatted[col] = formatted[col].astype(str)

        table = ax.table(
            cellText=formatted.values,
            colLabels=formatted.columns,
            loc="center",
            cellLoc="center",
            colLoc="center",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(7.0)
        table.scale(1, 1.18)

        for (row, _col), cell in table.get_celld().items():
            if row == 0:
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#D9EAF7")
            elif formatted.iloc[row - 1, 0] == "AVERAGE":
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#FFF2CC")

        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)


def add_summary_table(pdf: PdfPages, summary: pd.DataFrame) -> None:
    fig, ax = plt.subplots(figsize=(11.69, 8.27))
    ax.axis("off")
    ax.set_title("Course-Level Summary", fontsize=14, weight="bold", pad=14)

    table_df = summary.copy()
    for col in table_df.columns:
        if col not in {"Course", "Selected Models"}:
            table_df[col] = table_df[col].map(lambda value: f"{float(value):.2f}")

    table = ax.table(
        cellText=table_df.values,
        colLabels=table_df.columns,
        loc="center",
        cellLoc="center",
        colLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.0)
    table.scale(1, 1.45)
    for (row, _col), cell in table.get_celld().items():
        if row == 0:
            cell.set_text_props(weight="bold")
            cell.set_facecolor("#D9EAF7")

    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def write_report(
    path: Path,
    metrics_by_course: dict[str, pd.DataFrame],
    model_records: pd.DataFrame,
    summary: pd.DataFrame,
) -> None:
    with PdfPages(path) as pdf:
        add_text_page(
            pdf,
            "Assignment 02: Final Marks Prediction Report",
            [
                "Objective: Predict each afternoon-section student's final total score using only the assessment activities available after the 5th activity, then after the 6th activity, and so on until the second-last activity.",
                "Data split: For each course, the morning sheet is used as the training set and the afternoon sheet is used as the test set. The actual score is the Total column out of 100.",
                "Method: A separate model is trained for every course and activity horizon. The model predicts the student's remaining-score ratio, then adds it to the marks already earned by the known activities. Missing activity scores are treated as zero marks. The feature set includes raw scores, score ratios, weighted earned marks, observation flags, cumulative earned marks, and completion ratios.",
                "Modeling choices: Cloud Computing uses a Random Forest remaining-ratio model. ICT uses an Elastic Net remaining-ratio model. The training CV MAE column reports 5-fold cross-validation error on the morning section for the selected strategy at each horizon.",
                "Post-processing: Predicted totals are clipped to the valid score range of 0 to 100 and are not allowed to fall below the weighted score already earned by that activity horizon. A conservative engagement fallback is used for students with extremely low earned marks and mostly missing/zero known activities; for those cases the prediction is set to the current earned marks.",
                "Student metrics: For each student, absolute errors are calculated across all prediction horizons. The report table shows each student's average, maximum, and minimum absolute error, followed by an AVERAGE row for those three metrics.",
            ],
        )

        add_summary_table(pdf, summary)

        for course_name, metrics in metrics_by_course.items():
            add_dataframe_table(pdf, f"{course_name}: Student Error Metrics", metrics)

        add_dataframe_table(
            pdf,
            "Selected Model and MAE by Prediction Horizon",
            model_records,
            rows_per_page=30,
        )


def build_summary(
    all_metrics: dict[str, pd.DataFrame],
    model_records: pd.DataFrame,
    row_counts: dict[str, tuple[int, int]],
    activities: dict[str, int],
) -> pd.DataFrame:
    rows = []
    for course_name, metrics in all_metrics.items():
        average_row = metrics[metrics["Student ID"] == "AVERAGE"].iloc[0]
        course_models = model_records[model_records["Course"] == course_name]
        train_rows, test_rows = row_counts[course_name]
        rows.append(
            {
                "Course": course_name,
                "Training Students": train_rows,
                "Test Students": test_rows,
                "Activities": activities[course_name],
                "Prediction Horizons": len(course_models),
                "Mean Avg Abs Error": average_row["Average Absolute Error"],
                "Mean Max Abs Error": average_row["Max Absolute Error"],
                "Mean Min Abs Error": average_row["Min Absolute Error"],
            }
        )
    return pd.DataFrame(rows)


def run(data_dir: Path, output_dir: Path, random_state: int) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    predictions_path = output_dir / "Assignment_02_Predictions.xlsx"
    report_path = output_dir / "Assignment_02_Report.pdf"

    predictions_by_course: dict[str, pd.DataFrame] = {}
    metrics_by_course: dict[str, pd.DataFrame] = {}
    model_record_frames = []
    row_counts: dict[str, tuple[int, int]] = {}
    activities: dict[str, int] = {}

    for config in COURSES:
        train = load_course_dataset(data_dir, config, config.train_sheet)
        test = load_course_dataset(data_dir, config, config.test_sheet)
        predictions, metrics, model_records = predict_course(train, test, random_state)

        predictions_by_course[config.course_name] = predictions
        metrics_by_course[config.course_name] = metrics
        model_record_frames.append(model_records)
        row_counts[config.course_name] = (len(train.target), len(test.target))
        activities[config.course_name] = len(train.features.columns)

    all_model_records = pd.concat(model_record_frames, ignore_index=True)
    summary = build_summary(metrics_by_course, all_model_records, row_counts, activities)

    write_prediction_workbook(predictions_path, predictions_by_course)
    write_report(report_path, metrics_by_course, all_model_records, summary)

    print("Generated files:")
    print(f"- {predictions_path}")
    print(f"- {report_path}")
    print()
    print(summary.to_string(index=False))
    return predictions_path, report_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Assignment 02 predictions and report.")
    parser.add_argument("--data-dir", type=Path, default=Path("."), help="Directory containing the dataset files.")
    parser.add_argument("--output-dir", type=Path, default=Path("outputs"), help="Directory for generated outputs.")
    parser.add_argument("--random-state", type=int, default=42, help="Random seed for reproducibility.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    run(args.data_dir.resolve(), args.output_dir.resolve(), args.random_state)


if __name__ == "__main__":
    main()
